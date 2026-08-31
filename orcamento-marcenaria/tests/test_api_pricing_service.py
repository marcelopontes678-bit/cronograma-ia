import math
from datetime import datetime, timezone

import openpyxl
import pytest
from openpyxl.styles import Font

from api.schemas.extracao import (
    Ambiente,
    AuditoriaVisual,
    Componentes,
    Dimensoes,
    EspecificacoesMateriais,
    ExtracaoResultado,
    Modulo,
    OrigemModulo,
    StatusExtracao,
)
from api.services.pricing_service import PrecificacaoInvalidaError, gerar_orcamento


def _auditoria():
    return AuditoriaVisual(pagina_pdf=5, bounding_box=[100, 100, 200, 200])


def _materiais(caixaria="MDF Cinza Cobalto Berneck"):
    return EspecificacoesMateriais(
        caixaria=caixaria, frente=caixaria, fundo="MDF Branco",
        metodo_uniao="minifix", fixacao_fundo="encaixado_em_rebaixo",
    )


def _tabela_precos_teste(caminho, preco_chapa=450.0, preco_fita=28.0):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "REFERENCE (codigo Promob)", "Codigo Interno", "Descricao", "Categoria", "Espessura (mm)", "Unidade",
        "Preco Unitario UN (R$)", "Preco Chapa Fechada (R$)", "Preco Fita de Borda (R$/m)",
        "Fornecedor", "Data Atualizacao", "Observacoes",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    ws.cell(row=5, column=1, value="MDF Cinza Cobalto Berneck")
    ws.cell(row=5, column=6, value="M2")
    ws.cell(row=5, column=8, value=preco_chapa)
    ws.cell(row=5, column=9, value=preco_fita)
    wb.save(caminho)
    return caminho


@pytest.fixture
def resultado_confirmado():
    modulo = Modulo(
        id="mod_0001", nome="Armario superior espelhado",
        dimensoes=Dimensoes(largura_mm=930, altura_mm=1040, profundidade_mm=150),
        componentes=Componentes(portas=1, gavetas=0),
        especificacoes_materiais=_materiais(),
        auditoria_visual=_auditoria(),
        confianca=0.9, origem=OrigemModulo.CONFIRMADO_HUMANO,
    )
    return ExtracaoResultado(
        job_id="job_pricing_teste", arquivo_origem="banheiro.pdf", status=StatusExtracao.CONFIRMADO,
        ambientes=[Ambiente(nome_ambiente="Banheiro", modulos=[modulo])],
        avisos=[], criado_em=datetime.now(timezone.utc), atualizado_em=datetime.now(timezone.utc),
    )


class TestGerarOrcamentoSemFatorDeArea:
    def test_nunca_inventa_area_a_partir_da_dimensao_frontal(self, resultado_confirmado, tmp_path, caminho_config_precificacao):
        tabela = _tabela_precos_teste(tmp_path / "tabela.xlsx")
        resp, avisos = gerar_orcamento(
            resultado_confirmado, tabela, caminho_config_precificacao, faturamento_acumulado=100_000,
        )
        assert resp.custo_material_total == 0.0
        assert resp.total == 0.0
        assert any("fator_area_frontal_para_chapa nao informado" in a for a in avisos)


class TestGerarOrcamentoComFatorDeArea:
    def test_matematica_bate_com_calculo_manual(self, resultado_confirmado, tmp_path, caminho_config_precificacao):
        tabela = _tabela_precos_teste(tmp_path / "tabela.xlsx", preco_chapa=450.0, preco_fita=28.0)
        resp, avisos = gerar_orcamento(
            resultado_confirmado, tabela, caminho_config_precificacao, faturamento_acumulado=100_000,
            custo_hora_mao_de_obra=30, horas_estimadas=5, fator_area_frontal_para_chapa=2.2,
        )

        area_frontal = (930 / 1000) * (1040 / 1000)
        area_com_fator = area_frontal * 2.2
        area_com_perda = area_com_fator * 1.15  # PCT_PERDA_CORTE_PADRAO
        n_chapas = math.ceil(area_com_perda / (2.750 * 1.830))
        custo_chapa_esperado = n_chapas * 450.0

        perimetro = 2 * (1040 + 930) / 1000
        custo_fita_esperado = perimetro * 28.0

        assert resp.custo_material_total == pytest.approx(custo_chapa_esperado + custo_fita_esperado)
        assert resp.custo_mao_de_obra == 150.0
        assert resp.total == pytest.approx(resp.preco_venda_material + 150.0)
        assert any("contagem de ferragens" in a for a in avisos)  # portas/gavetas ainda nao viram ferragem

    def test_modulo_sem_dimensoes_fica_pendente(self, tmp_path, caminho_config_precificacao):
        modulo_sem_dim = Modulo(
            id="mod_0002", nome="Prateleira solta",
            especificacoes_materiais=_materiais(caixaria="MDF Branco"),
            auditoria_visual=_auditoria(),
            confianca=0.4, origem=OrigemModulo.VISION_AUTOMATICO,
        )
        resultado = ExtracaoResultado(
            job_id="j2", arquivo_origem="x.pdf", status=StatusExtracao.CONFIRMADO,
            ambientes=[Ambiente(nome_ambiente="Banheiro", modulos=[modulo_sem_dim])],
            criado_em=datetime.now(timezone.utc), atualizado_em=datetime.now(timezone.utc),
        )
        tabela = _tabela_precos_teste(tmp_path / "tabela.xlsx")
        resp, avisos = gerar_orcamento(
            resultado, tabela, caminho_config_precificacao, faturamento_acumulado=100_000,
            fator_area_frontal_para_chapa=2.2,
        )
        assert resp.custo_material_total == 0.0
        assert any("sem largura/altura extraidas" in a for a in avisos)


class TestStatusNaoConfirmado:
    def test_rejeita_extracao_nao_confirmada(self, resultado_confirmado, tmp_path, caminho_config_precificacao):
        resultado_nao_confirmado = resultado_confirmado.model_copy(update={"status": StatusExtracao.AGUARDANDO_REVISAO})
        tabela = _tabela_precos_teste(tmp_path / "tabela.xlsx")
        with pytest.raises(PrecificacaoInvalidaError):
            gerar_orcamento(resultado_nao_confirmado, tabela, caminho_config_precificacao, faturamento_acumulado=100_000)
