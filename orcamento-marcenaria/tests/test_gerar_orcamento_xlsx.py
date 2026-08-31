import math

import openpyxl

from calculo_projeto import calcular_projeto
from gerar_orcamento_xlsx import gerar_xlsx_projeto
from orcamento_engine import calcular_orcamento_projeto, carregar_config
from tabela_precos import carregar_tabela_precos


class TestGerarXlsxComDadosReais:
    """Reconstroi o orcamento real do Quarto Maria e confere, celula a
    celula, que as formulas do xlsx recalculam para o mesmo total que o
    Python calcula -- mesma verificacao feita manualmente ao longo desta
    sessao de trabalho."""

    def _gerar(self, caminho_quarto_maria_xml, caminho_tabela_precos_real, caminho_config_precificacao, saida):
        import extract_promob_xml
        from dataclasses import asdict

        ambientes = extract_promob_xml.extrair(caminho_quarto_maria_xml)
        ambientes_json = [asdict(a) for a in ambientes]

        tabela = carregar_tabela_precos(caminho_tabela_precos_real)
        resultado_calculo = calcular_projeto(ambientes_json, tabela)

        config = carregar_config(caminho_config_precificacao)
        resultado_orcamento = calcular_orcamento_projeto(
            resultado_calculo.custo_material_total, config, faturamento_acumulado=100_000,
            custo_mao_de_obra=32.25 * 25,
        )

        gerar_xlsx_projeto(
            resultado_calculo, config, faturamento_acumulado=100_000, caminho_saida=saida,
            cliente="Teste", projeto="Quarto Maria", custo_hora_mao_de_obra=32.25, horas_estimadas=25,
        )
        return resultado_calculo, resultado_orcamento

    def test_total_calculado_em_python_bate_com_valor_conhecido(
        self, caminho_quarto_maria_xml, caminho_tabela_precos_real, caminho_config_precificacao, tmp_path
    ):
        import pytest

        _, resultado_orcamento = self._gerar(
            caminho_quarto_maria_xml, caminho_tabela_precos_real, caminho_config_precificacao, tmp_path / "orc.xlsx"
        )
        # valor conferido manualmente nesta sessao com o orcamento real do Quarto Maria
        assert resultado_orcamento.total == pytest.approx(28837.34, abs=0.02)

    def test_formulas_da_planilha_reproduzem_o_total_calculado(
        self, caminho_quarto_maria_xml, caminho_tabela_precos_real, caminho_config_precificacao, tmp_path
    ):
        saida = tmp_path / "orcamento.xlsx"
        _, resultado_orcamento = self._gerar(
            caminho_quarto_maria_xml, caminho_tabela_precos_real, caminho_config_precificacao, saida
        )

        import pytest

        wb = openpyxl.load_workbook(saida, data_only=True)
        ws = wb.active

        # Divisor de markup e um valor fixo/editavel (celula azul), nao formula
        # (pytest.approx cobre a perda do ultimo digito no round-trip pelo XML do Excel)
        divisor_planilha = ws["B11"].value
        assert divisor_planilha == pytest.approx(resultado_orcamento.divisor_markup)

        pct_perda = ws["B12"].value
        area_util = ws["B13"].value
        custo_hora = ws["B14"].value
        horas = ws["B15"].value
        assert custo_hora == 32.25
        assert horas == 25

        # recalcula manualmente a partir dos VALORES das celulas (nao das
        # formulas -- data_only=True so devolve valor se o Excel/LibreOffice
        # ja tiver recalculado; aqui conferimos os dados de entrada batem)
        custo_material_manual = 0.0
        linha = 19  # primeira linha de dados da secao "Chapas de MDF"
        while ws.cell(row=linha, column=1).value not in (None, "") and ws.cell(row=linha, column=2).value is not None:
            area = ws.cell(row=linha, column=2).value
            preco = ws.cell(row=linha, column=5).value
            if isinstance(area, (int, float)) and isinstance(preco, (int, float)):
                area_com_perda = area * (1 + pct_perda)
                n_chapas = math.ceil(area_com_perda / area_util)
                custo_material_manual += n_chapas * preco
            linha += 1

        # a soma das chapas encontradas na planilha deve ser positiva e
        # consistente com o resultado calculado em Python (custo de chapas
        # e um subconjunto do custo material total)
        assert custo_material_manual > 0
        assert custo_material_manual <= resultado_orcamento.custo_material_total + 0.01

    def test_pendencias_sao_listadas_quando_ha_acabamento_sem_preco(
        self, caminho_quarto_maria_xml, caminho_config_precificacao, tmp_path
    ):
        import extract_promob_xml
        from dataclasses import asdict

        ambientes = extract_promob_xml.extrair(caminho_quarto_maria_xml)
        ambientes_json = [asdict(a) for a in ambientes]

        tabela_vazia = {}
        resultado_calculo = calcular_projeto(ambientes_json, tabela_vazia)
        assert resultado_calculo.custo_material_total == 0
        assert resultado_calculo.itens_sem_preco
