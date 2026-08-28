"""Gera o orcamento final em .xlsx a partir de um ResultadoOrcamento.

Usa formulas (nao valores fixos calculados em Python) para que o preco de
venda, o total por modulo e o total geral recalculem se o usuario editar
o custo de material ou a mao de obra diretamente na planilha.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from orcamento_engine import ResultadoOrcamento

FONT = "Arial"
COR_HEADER_BG = "305496"
COR_HEADER_TEXTO = "FFFFFF"
COR_TOTAL_BG = "D9E1F2"
BORDA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
MOEDA = '$#,##0.00;($#,##0.00);"-"'


def gerar_xlsx(
    resultado: ResultadoOrcamento,
    caminho_saida: str | Path,
    cliente: str = "",
    projeto: str = "",
    faturamento_acumulado: float = 0.0,
) -> Path:
    caminho_saida = Path(caminho_saida)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orcamento"

    ws["A1"] = "Orcamento de Marcenaria"
    ws["A1"].font = Font(name=FONT, bold=True, size=16)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Cliente: {cliente or '-'}"
    ws["A2"].font = Font(name=FONT, size=11)
    ws["A3"] = f"Projeto: {projeto or '-'}"
    ws["A3"].font = Font(name=FONT, size=11)

    linha_params = 5
    ws.cell(row=linha_params, column=1, value="Divisor de Markup").font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=linha_params, column=2, value=resultado.divisor_markup).number_format = "0.0000"
    ws.cell(row=linha_params + 1, column=1, value="% Comissao de Vendas Aplicada").font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=linha_params + 1, column=2, value=resultado.pct_comissao_vendas_aplicada).number_format = "0.0%"
    ws.cell(row=linha_params + 2, column=1, value="Faturamento Acumulado Informado (R$)").font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=linha_params + 2, column=2, value=faturamento_acumulado).number_format = MOEDA

    linha_cabecalho = linha_params + 4
    headers = ["Ambiente / Modulo", "Custo Material (R$)", "Divisor Markup", "Preco Venda Material (R$)", "Mao de Obra (R$)", "Preco Final (R$)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=linha_cabecalho, column=col, value=h)
        c.font = Font(name=FONT, bold=True, color=COR_HEADER_TEXTO, size=11)
        c.fill = PatternFill(start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA

    celula_divisor = f"$B${linha_params}"
    linha_atual = linha_cabecalho + 1
    primeira_linha_dados = linha_atual
    for modulo in resultado.modulos:
        ws.cell(row=linha_atual, column=1, value=modulo.nome).border = BORDA
        cel_custo = ws.cell(row=linha_atual, column=2, value=modulo.custo_material)
        cel_custo.number_format = MOEDA
        cel_custo.border = BORDA

        cel_div = ws.cell(row=linha_atual, column=3, value=f"={celula_divisor}")
        cel_div.number_format = "0.0000"
        cel_div.border = BORDA

        col_custo_letra = get_column_letter(2)
        col_div_letra = get_column_letter(3)
        cel_venda = ws.cell(row=linha_atual, column=4, value=f"={col_custo_letra}{linha_atual}*{col_div_letra}{linha_atual}")
        cel_venda.number_format = MOEDA
        cel_venda.border = BORDA

        cel_mao_obra = ws.cell(row=linha_atual, column=5, value=modulo.custo_mao_de_obra)
        cel_mao_obra.number_format = MOEDA
        cel_mao_obra.border = BORDA

        col_venda_letra = get_column_letter(4)
        col_mao_obra_letra = get_column_letter(5)
        cel_final = ws.cell(row=linha_atual, column=6, value=f"={col_venda_letra}{linha_atual}+{col_mao_obra_letra}{linha_atual}")
        cel_final.number_format = MOEDA
        cel_final.border = BORDA

        linha_atual += 1

    ultima_linha_dados = linha_atual - 1

    ws.cell(row=linha_atual, column=1, value="TOTAL GERAL").font = Font(name=FONT, bold=True)
    ws.cell(row=linha_atual, column=1).fill = PatternFill(start_color=COR_TOTAL_BG, end_color=COR_TOTAL_BG, fill_type="solid")
    for col in range(2, 7):
        letra = get_column_letter(col)
        c = ws.cell(row=linha_atual, column=col, value=f"=SUM({letra}{primeira_linha_dados}:{letra}{ultima_linha_dados})")
        c.number_format = MOEDA
        c.font = Font(name=FONT, bold=True)
        c.fill = PatternFill(start_color=COR_TOTAL_BG, end_color=COR_TOTAL_BG, fill_type="solid")
        c.border = BORDA

    larguras = [42, 18, 14, 20, 16, 16]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1).coordinate

    wb.save(caminho_saida)
    return caminho_saida


FONTE_INPUT = Font(name=FONT, color="0000FF")  # celulas editaveis (convencao: azul = input)


def gerar_xlsx_projeto(
    resultado_calculo,  # calculo_projeto.ResultadoCalculoProjeto
    config: dict,
    faturamento_acumulado: float,
    caminho_saida: str | Path,
    cliente: str = "",
    projeto: str = "",
    custo_hora_mao_de_obra: float = 0.0,
    horas_estimadas: float = 0.0,
    pct_perda_corte: float = 0.15,
    area_util_chapa_m2: float = 2.750 * 1.830,
) -> Path:
    """Gera o orcamento no modelo por chapa fechada + fita de borda + ferragens,
    com TUDO ligado por formulas: mudar um preco, a % de perda, a area da
    chapa, o custo-hora ou as horas estimadas recalcula o orcamento inteiro
    sozinho no Excel, sem precisar rodar o script de novo. Celulas em AZUL
    sao as editaveis; o resto e formula."""
    from orcamento_engine import calcular_divisor_markup, pct_comissao_vendas

    caminho_saida = Path(caminho_saida)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orcamento"

    ws["A1"] = "Orcamento de Marcenaria"
    ws["A1"].font = Font(name=FONT, bold=True, size=16)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Cliente: {cliente or '-'}"
    ws["A3"] = f"Projeto: {projeto or '-'}"
    ws["A2"].font = Font(name=FONT, size=11)
    ws["A3"].font = Font(name=FONT, size=11)

    # --- Parametros editaveis (azul) ---
    m = config["markup"]
    divisor_atual = calcular_divisor_markup(config, faturamento_acumulado)
    pct_comissao_atual = pct_comissao_vendas(config, faturamento_acumulado)

    linha = 5
    ws.cell(row=linha, column=1, value="% Custo Fixo").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=m["pct_custo_fixo"]); c.number_format = "0.00%"; c.font = FONTE_INPUT
    linha_custo_fixo = linha
    linha += 1
    ws.cell(row=linha, column=1, value="% Impostos").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=m["pct_impostos"]); c.number_format = "0.00%"; c.font = FONTE_INPUT
    linha_impostos = linha
    linha += 1
    ws.cell(row=linha, column=1, value="% Comissao Fabrica").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=m["pct_comissao_fabrica"]); c.number_format = "0.00%"; c.font = FONTE_INPUT
    linha_comissao_fabrica = linha
    linha += 1
    ws.cell(row=linha, column=1, value="% Comissao Vendas (faixa aplicada)").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=pct_comissao_atual); c.number_format = "0.00%"; c.font = FONTE_INPUT
    linha_comissao_vendas = linha
    linha += 1
    ws.cell(row=linha, column=1, value="% Lucro").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=m["pct_lucro"]); c.number_format = "0.00%"; c.font = FONTE_INPUT
    linha_lucro = linha
    linha += 1
    ws.cell(row=linha, column=1, value="Divisor de Markup").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(
        row=linha, column=2,
        value=f"=1/(1-(B{linha_custo_fixo}+B{linha_impostos}+B{linha_comissao_fabrica}+B{linha_comissao_vendas}+B{linha_lucro}))",
    )
    c.number_format = "0.0000"
    celula_divisor = f"$B${linha}"
    linha += 1
    ws.cell(row=linha, column=1, value="% Perda de Corte").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=pct_perda_corte); c.number_format = "0.0%"; c.font = FONTE_INPUT
    celula_pct_perda = f"$B${linha}"
    linha += 1
    ws.cell(row=linha, column=1, value="Area Util da Chapa (m2)").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=area_util_chapa_m2); c.number_format = "0.0000"; c.font = FONTE_INPUT
    celula_area_util = f"$B${linha}"
    linha += 1
    ws.cell(row=linha, column=1, value="Custo Hora Mao de Obra (R$)").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=custo_hora_mao_de_obra); c.number_format = MOEDA; c.font = FONTE_INPUT
    celula_custo_hora = f"$B${linha}"
    linha += 1
    ws.cell(row=linha, column=1, value="Horas Estimadas").font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=linha, column=2, value=horas_estimadas); c.number_format = "0.0"; c.font = FONTE_INPUT
    celula_horas = f"$B${linha}"

    def escrever_secao(titulo, headers, linha_inicio):
        l = linha_inicio
        ws.cell(row=l, column=1, value=titulo).font = Font(name=FONT, bold=True, size=12)
        l += 1
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=l, column=col, value=h)
            c.font = Font(name=FONT, bold=True, color=COR_HEADER_TEXTO)
            c.fill = PatternFill(start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid")
            c.border = BORDA
            c.alignment = Alignment(horizontal="center")
        return l + 1

    linha += 2

    # --- Chapas de MDF: Area c/ Perda, Num Chapas e Custo viram formula;
    # Preco/Chapa fica editavel (azul). ---
    linha_inicio_chapas = escrever_secao(
        "Chapas de MDF",
        ["Acabamento", "Area Total (m2)", "Area c/ Perda (m2)", "Num Chapas", "Preco/Chapa (R$)", "Custo (R$)"],
        linha,
    )
    p1 = linha_inicio_chapas
    for i, c_dado in enumerate(resultado_calculo.chapas):
        r = linha_inicio_chapas + i
        ws.cell(row=r, column=1, value=c_dado.acabamento).border = BORDA
        cel_area = ws.cell(row=r, column=2, value=c_dado.area_total_m2); cel_area.border = BORDA; cel_area.number_format = "0.000"
        cel_perda = ws.cell(row=r, column=3, value=f"=B{r}*(1+{celula_pct_perda})"); cel_perda.border = BORDA; cel_perda.number_format = "0.000"
        cel_num = ws.cell(row=r, column=4, value=f"=ROUNDUP(C{r}/{celula_area_util},0)"); cel_num.border = BORDA
        cel_preco = ws.cell(row=r, column=5, value=c_dado.preco_chapa); cel_preco.border = BORDA; cel_preco.number_format = MOEDA; cel_preco.font = FONTE_INPUT
        cel_custo = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cel_custo.border = BORDA; cel_custo.number_format = MOEDA
    u1 = linha_inicio_chapas + len(resultado_calculo.chapas) - 1
    linha = u1 + 2 if resultado_calculo.chapas else linha_inicio_chapas + 1

    # --- Fita de Borda: Custo vira formula; Preco/Metro fica editavel. ---
    linha_inicio_fitas = escrever_secao(
        "Fita de Borda",
        ["Acabamento", "Metros Total", "Preco/Metro (R$)", "Custo (R$)"],
        linha,
    )
    p2 = linha_inicio_fitas
    for i, f_dado in enumerate(resultado_calculo.fitas):
        r = linha_inicio_fitas + i
        ws.cell(row=r, column=1, value=f_dado.acabamento).border = BORDA
        cel_metros = ws.cell(row=r, column=2, value=f_dado.metros_total); cel_metros.border = BORDA; cel_metros.number_format = "0.000"
        cel_preco = ws.cell(row=r, column=3, value=f_dado.preco_fita_metro); cel_preco.border = BORDA; cel_preco.number_format = MOEDA; cel_preco.font = FONTE_INPUT
        cel_custo = ws.cell(row=r, column=4, value=f"=B{r}*C{r}"); cel_custo.border = BORDA; cel_custo.number_format = MOEDA
    u2 = linha_inicio_fitas + len(resultado_calculo.fitas) - 1
    linha = u2 + 2 if resultado_calculo.fitas else linha_inicio_fitas + 1

    # --- Ferragens / Componentes: Custo vira formula; Preco Unitario editavel. ---
    linha_inicio_ferragens = escrever_secao(
        "Ferragens / Componentes",
        ["Descricao", "Referencia", "Preco Unitario (R$)", "Quantidade", "Custo (R$)"],
        linha,
    )
    p3 = linha_inicio_ferragens
    for i, fe in enumerate(resultado_calculo.ferragens):
        r = linha_inicio_ferragens + i
        ws.cell(row=r, column=1, value=fe.descricao).border = BORDA
        ws.cell(row=r, column=2, value=fe.reference).border = BORDA
        cel_preco = ws.cell(row=r, column=3, value=fe.preco_unitario); cel_preco.border = BORDA; cel_preco.number_format = MOEDA; cel_preco.font = FONTE_INPUT
        cel_qtd = ws.cell(row=r, column=4, value=fe.quantidade); cel_qtd.border = BORDA
        cel_custo = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cel_custo.border = BORDA; cel_custo.number_format = MOEDA
    u3 = linha_inicio_ferragens + len(resultado_calculo.ferragens) - 1
    linha = u3 + 2 if resultado_calculo.ferragens else linha_inicio_ferragens + 1

    ws.cell(row=linha, column=1, value="Custo Material Total").font = Font(name=FONT, bold=True)
    partes = []
    if resultado_calculo.chapas:
        partes.append(f"SUM(F{p1}:F{u1})")
    if resultado_calculo.fitas:
        partes.append(f"SUM(D{p2}:D{u2})")
    if resultado_calculo.ferragens:
        partes.append(f"SUM(E{p3}:E{u3})")
    f_custo = "=" + "+".join(partes) if partes else 0
    ws.cell(row=linha, column=2, value=f_custo).number_format = MOEDA
    linha_custo_material = linha

    linha += 1
    ws.cell(row=linha, column=1, value="Preco Venda Material (x Markup)").font = Font(name=FONT, bold=True)
    ws.cell(row=linha, column=2, value=f"=B{linha_custo_material}*{celula_divisor}").number_format = MOEDA
    linha_venda = linha

    linha += 1
    ws.cell(row=linha, column=1, value="Mao de Obra (Custo Hora x Horas)").font = Font(name=FONT, bold=True)
    ws.cell(row=linha, column=2, value=f"={celula_custo_hora}*{celula_horas}").number_format = MOEDA
    linha_mao_obra = linha

    linha += 1
    ws.cell(row=linha, column=1, value="TOTAL GERAL").font = Font(name=FONT, bold=True, size=12)
    ws.cell(row=linha, column=1).fill = PatternFill(start_color=COR_TOTAL_BG, end_color=COR_TOTAL_BG, fill_type="solid")
    c_total = ws.cell(row=linha, column=2, value=f"=B{linha_venda}+B{linha_mao_obra}")
    c_total.number_format = MOEDA
    c_total.font = Font(name=FONT, bold=True, size=12)
    c_total.fill = PatternFill(start_color=COR_TOTAL_BG, end_color=COR_TOTAL_BG, fill_type="solid")

    if resultado_calculo.itens_sem_preco:
        linha += 2
        ws.cell(row=linha, column=1, value=f"Pendencias sem preco na tabela ({len(resultado_calculo.itens_sem_preco)}) -- NAO incluidas no total:").font = Font(name=FONT, italic=True, color="C00000")
        linha += 1
        for ref, desc, motivo in resultado_calculo.itens_sem_preco:
            ws.cell(row=linha, column=1, value=f"  {motivo}: {ref} ({desc})").font = Font(name=FONT, italic=True, size=9, color="666666")
            linha += 1

    linha += 2
    ws.cell(row=linha, column=1, value="Legenda: celulas em AZUL sao editaveis (precos, %, horas). O resto recalcula sozinho.").font = Font(name=FONT, italic=True, size=9, color="0000FF")

    for i, w in enumerate([42, 18, 18, 14, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(caminho_saida)
    return caminho_saida


if __name__ == "__main__":
    import argparse
    import json
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from calculo_projeto import calcular_projeto
    from orcamento_engine import calcular_orcamento_projeto, carregar_config
    from tabela_precos import carregar_tabela_precos

    parser = argparse.ArgumentParser(description="Gera orcamento final .xlsx (chapa/fita/ferragem) a partir de uma extracao Promob + tabela de precos.")
    parser.add_argument("arquivo_extracao_json", help="JSON gerado por extract_promob_xml.py")
    parser.add_argument("--tabela-precos", default=str(Path(__file__).resolve().parent.parent / "config" / "tabela_precos_referencia.xlsx"))
    parser.add_argument("--config", default=str(Path(__file__).resolve().parent.parent / "config" / "precificacao.json"))
    parser.add_argument("--faturamento-acumulado", type=float, required=True)
    parser.add_argument("--custo-hora", type=float, default=0.0, help="Custo-hora da mao de obra (R$/h)")
    parser.add_argument("--horas-estimadas", type=float, default=0.0, help="Horas estimadas de mao de obra")
    parser.add_argument("--cliente", default="")
    parser.add_argument("--saida", default=str(Path(__file__).resolve().parent.parent / "output" / "orcamento_final.xlsx"))
    args = parser.parse_args()

    data = json.load(open(args.arquivo_extracao_json, encoding="utf-8"))
    tabela = carregar_tabela_precos(args.tabela_precos)
    resultado_calculo = calcular_projeto(data, tabela)

    config = carregar_config(args.config)
    custo_mao_de_obra = args.custo_hora * args.horas_estimadas
    resultado_orcamento = calcular_orcamento_projeto(
        resultado_calculo.custo_material_total,
        config,
        args.faturamento_acumulado,
        custo_mao_de_obra=custo_mao_de_obra,
    )

    caminho = gerar_xlsx_projeto(
        resultado_calculo,
        config,
        args.faturamento_acumulado,
        args.saida,
        cliente=args.cliente,
        projeto=data[0]["nome"] if data else "",
        custo_hora_mao_de_obra=args.custo_hora,
        horas_estimadas=args.horas_estimadas,
    )
    print(f"Orcamento xlsx gerado: {caminho}")
    print(f"Custo material total: R${resultado_calculo.custo_material_total:.2f}")
    print(f"TOTAL (calculado em Python, para conferencia): R${resultado_orcamento.total:.2f}")
    if resultado_calculo.itens_sem_preco:
        print(f"ATENCAO: {len(resultado_calculo.itens_sem_preco)} pendencias sem preco na tabela (nao incluidas no total)")
