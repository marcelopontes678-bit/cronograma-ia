"""Carrega a tabela de precos por referencia Promob (config/tabela_precos_referencia.xlsx).

Modelo de precificacao:
- Chapas de MDF (itens com unidade M2): precificadas por CHAPA FECHADA
  (preco_chapa_fechada, R$/chapa de 2750x1830mm) e FITA DE BORDA
  (preco_fita_metro, R$/m) -- agrupadas por ACABAMENTO (espessura +
  nome do material extraidos do REFERENCE), NAO pelo REFERENCE completo.
  O REFERENCE do Promob varia por tipo de peca (Base, Lateral, Fundo...)
  mesmo quando a peca e cortada da mesma chapa/acabamento -- ver
  `chave_acabamento()` e engine/calculo_projeto.py.
- Ferragens/componentes (itens com unidade UN): preco por unidade
  (preco_unitario_un) x quantidade/repeticao. Um item UN cuja
  Observacoes contenha "entra na chapa" e tratado como custo ZERO
  explicito (informado pelo usuario -- o material dele ja esta contado
  como parte de uma chapa MDF em outro item), nao como pendencia.

Nao inventa preco para acabamento/referencia sem dado na tabela: fica
sinalizado como pendente, nunca com custo zero ou estimado (exceto o
caso explicito "entra na chapa" acima, que e informacao do usuario).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class PrecoReferencia:
    reference: str
    codigo_interno: str
    descricao: str
    categoria: str
    espessura_mm: float | None
    unidade: str
    preco_unitario_un: float | None       # para itens UN (ferragens/componentes)
    preco_chapa_fechada: float | None     # para itens M2 (chapa MDF, preco da chapa 2750x1830mm inteira)
    preco_fita_metro: float | None        # para itens M2 (fita de borda do mesmo acabamento, R$/m)
    fornecedor: str
    observacoes: str = ""


def chave_acabamento(reference: str) -> tuple[int, str] | None:
    """Extrai (espessura_mm, nome_do_acabamento) do REFERENCE Promob.
    Padrao tipico: <codigo>.<codigo2>.<espessura>.<nome_do_acabamento>[.MDF|.Aglom]
    Ex: '2.2008.18.Duratex.Essencial.Rosa Infinito.MDF' -> (18, 'Duratex.Essencial.Rosa Infinito')
    Retorna None se o REFERENCE nao seguir esse padrao (ex: codigos de ferragem)."""
    m = re.match(r"^\d+\.\d+\.(\d+)\.(.+)$", reference or "")
    if not m:
        return None
    espessura, resto = m.groups()
    resto = re.sub(r"\.(MDF|Aglom)$", "", resto)
    return (int(espessura), resto)


# Colunas da planilha (config/tabela_precos_referencia.xlsx), na ordem:
# REFERENCE, Codigo Interno, Descricao, Categoria, Espessura (mm), Unidade,
# Preco Unitario UN (R$), Preco Chapa Fechada (R$), Preco Fita de Borda (R$/m),
# Fornecedor, Data Atualizacao, Observacoes
def carregar_tabela_precos(caminho_xlsx: str | Path) -> dict[str, PrecoReferencia]:
    caminho_xlsx = Path(caminho_xlsx)
    if not caminho_xlsx.exists():
        raise FileNotFoundError(f"Tabela de precos nao encontrada: {caminho_xlsx}")

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb.active

    tabela: dict[str, PrecoReferencia] = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        reference = row[0].value
        if not reference:
            continue

        preco_un = row[6].value
        preco_chapa = row[7].value
        preco_fita = row[8].value
        espessura = row[4].value
        observacoes = row[11].value if len(row) > 11 else None

        if preco_un is None and preco_chapa is None and preco_fita is None and not observacoes:
            continue

        tabela[reference] = PrecoReferencia(
            reference=reference,
            codigo_interno=row[1].value or "",
            descricao=row[2].value or "",
            categoria=row[3].value or "",
            espessura_mm=float(espessura) if espessura is not None else None,
            unidade=row[5].value or "",
            preco_unitario_un=float(preco_un) if preco_un is not None else None,
            preco_chapa_fechada=float(preco_chapa) if preco_chapa is not None else None,
            preco_fita_metro=float(preco_fita) if preco_fita is not None else None,
            fornecedor=row[9].value or "",
            observacoes=observacoes or "",
        )
    return tabela


def indexar_precos_por_acabamento(tabela: dict[str, PrecoReferencia]) -> dict[tuple[int, str], PrecoReferencia]:
    """Agrupa as entradas da tabela por (espessura, nome_acabamento), pra casar
    pecas MDF de REFERENCE diferente que vem da mesma chapa/acabamento.

    Mesmo fallback usado por calculo_projeto.py ao agrupar itens: quando o
    REFERENCE nao segue o padrao numerico do Promob (ex: nome de material
    em texto livre vindo de uma extracao por Vision), usa o proprio texto
    como chave (espessura=0) em vez de descartar a linha -- senao uma
    linha de preco com REFERENCE em texto livre nunca seria encontrada."""
    indice: dict[tuple[int, str], PrecoReferencia] = {}
    for preco_ref in tabela.values():
        chave = chave_acabamento(preco_ref.reference) or (0, preco_ref.reference)
        existente = indice.get(chave)
        # prefere uma entrada que ja tenha preco de chapa/fita preenchido
        if existente is None or (existente.preco_chapa_fechada is None and preco_ref.preco_chapa_fechada is not None):
            indice[chave] = preco_ref
    return indice


def calcular_custo_item_ferragem(item: dict, tabela: dict[str, PrecoReferencia]) -> tuple[float | None, str]:
    """Para itens de unidade UN (ferragens/componentes). Retorna (custo, status).
    custo e None quando a referencia nao tem preco cadastrado nem esta marcada
    como incluida em outra chapa."""
    ref = item.get("reference")
    preco_ref = tabela.get(ref)

    if preco_ref is not None and preco_ref.preco_unitario_un is not None:
        repeticao = item.get("repeticao", 1)
        return preco_ref.preco_unitario_un * repeticao, "OK"

    if preco_ref is not None and "entra na chapa" in preco_ref.observacoes.lower():
        return 0.0, "OK_INCLUIDO_NA_CHAPA"

    return None, "SEM_PRECO_NA_TABELA"
